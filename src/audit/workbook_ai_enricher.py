from __future__ import annotations

import argparse
import json
import os
import re
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
AI_HEADERS = [
    "AI Reviewed Status",
    "AI Reviewed Confidence",
    "Final Status",
    "Final Confidence",
    "AI Evidence Summary",
    "AI Key Insight",
    "AI UX Impact",
    "AI Recommended Fix",
    "AI Severity Hint",
    "AI Needs Human Review",
    "AI Contradiction Flag",
    "Recurrence Count",
    "Systemic Issue",
    "Pattern Label",
    "Decision Source",
    "AI Review Note",
]


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return str(value)


def _safe_int(value: Any) -> Optional[int]:
    try:
        if value in ("", None):
            return None
        return int(float(value))
    except Exception:
        return None


def _load_enriched_rows(enriched_json_path: str) -> Dict[str, Any]:
    with open(enriched_json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _apply_header_style(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autofit_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_idx = col_cells[0].column
        for cell in col_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            except Exception:
                continue
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 50)


def _normalize_comparable_text(value: Any) -> str:
    return re.sub(r"\s+", " ", _safe_str(value)).strip().casefold()


def _find_header_columns(ws) -> Dict[str, int]:
    header_columns: Dict[str, int] = {}
    for column in range(1, ws.max_column + 1):
        header = _normalize_comparable_text(ws.cell(row=1, column=column).value)
        if header:
            header_columns[header] = column
    return header_columns


def _has_existing_ai_columns(ws) -> bool:
    header_columns = _find_header_columns(ws)
    return any(_normalize_comparable_text(header) in header_columns for header in AI_HEADERS)


def _json_source_row_number(row: Dict[str, Any]) -> Optional[int]:
    raw = row.get("raw")
    if isinstance(raw, dict):
        source_row = _safe_int(raw.get("row"))
        if source_row is not None:
            return source_row

    return _safe_int(row.get("_source_row"))


def _validate_row_alignment(ws, rows: List[Dict[str, Any]]) -> None:
    header_columns = _find_header_columns(ws)
    required_headers = ["sheet", "row", "criterion"]
    missing_headers = [header for header in required_headers if header not in header_columns]
    if missing_headers:
        raise ValueError(
            "Generated Evidence sheet is missing required columns for AI alignment validation: "
            + ", ".join(missing_headers)
        )

    for excel_row, json_row in enumerate(rows, start=2):
        workbook_sheet = _normalize_comparable_text(ws.cell(row=excel_row, column=header_columns["sheet"]).value)
        workbook_row = _safe_int(ws.cell(row=excel_row, column=header_columns["row"]).value)
        workbook_criterion = _normalize_comparable_text(
            ws.cell(row=excel_row, column=header_columns["criterion"]).value
        )

        expected_sheet = _normalize_comparable_text(json_row.get("_sheet_name"))
        expected_row = _json_source_row_number(json_row)
        expected_criterion = _normalize_comparable_text(json_row.get("criterion"))

        mismatches = []
        if expected_sheet and workbook_sheet != expected_sheet:
            mismatches.append(f"sheet workbook='{workbook_sheet}' json='{expected_sheet}'")
        if expected_row is not None and workbook_row != expected_row:
            mismatches.append(f"row workbook='{workbook_row}' json='{expected_row}'")
        if expected_criterion and workbook_criterion != expected_criterion:
            mismatches.append("criterion differs")

        if mismatches:
            raise ValueError(
                f"Row alignment mismatch at Generated Evidence row {excel_row}: "
                + "; ".join(mismatches)
                + ". Use the workbook exported from the same checks JSON before AI enrichment."
            )


def _append_ai_columns(ws, rows: List[Dict[str, Any]]) -> None:
    start_col = ws.max_column + 1

    for offset, header in enumerate(AI_HEADERS):
        cell = ws.cell(row=1, column=start_col + offset, value=header)
        _apply_header_style(cell)

    excel_row = 2
    for row in rows:
        ws.cell(row=excel_row, column=start_col + 0, value=_safe_str(row.get("reviewed_status")))
        ws.cell(row=excel_row, column=start_col + 1, value=row.get("reviewed_confidence"))
        ws.cell(row=excel_row, column=start_col + 2, value=_safe_str(row.get("final_status")))
        ws.cell(row=excel_row, column=start_col + 3, value=row.get("final_confidence"))
        ws.cell(row=excel_row, column=start_col + 4, value=_safe_str(row.get("evidence_summary")))
        ws.cell(row=excel_row, column=start_col + 5, value=_safe_str(row.get("key_insight")))
        ws.cell(row=excel_row, column=start_col + 6, value=_safe_str(row.get("ux_impact")))
        ws.cell(row=excel_row, column=start_col + 7, value=_safe_str(row.get("recommended_fix")))
        ws.cell(row=excel_row, column=start_col + 8, value=_safe_str(row.get("severity_hint")))
        ws.cell(row=excel_row, column=start_col + 9, value="Yes" if row.get("needs_human_review") else "No")
        ws.cell(row=excel_row, column=start_col + 10, value="Yes" if row.get("contradiction_flag") else "No")
        ws.cell(row=excel_row, column=start_col + 11, value=row.get("recurrence_count", 1))
        ws.cell(row=excel_row, column=start_col + 12, value="Yes" if row.get("systemic_issue") else "No")
        ws.cell(row=excel_row, column=start_col + 13, value=_safe_str(row.get("group_pattern_label")))
        ws.cell(row=excel_row, column=start_col + 14, value=_safe_str(row.get("decision_source")))
        ws.cell(row=excel_row, column=start_col + 15, value=_safe_str(row.get("review_note")))
        excel_row += 1

    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def _write_summary_sheet(wb, recurrence_summary: List[Dict[str, Any]]) -> None:
    sheet_name = "AI Insights Summary"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)

    headers = [
        "Sheet",
        "Pattern Label",
        "Affected Rows",
        "Affected Pages",
        "Systemic Issue",
        "Highest Severity",
        "Representative Insight",
        "Recommended Action",
        "Example URLs",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _apply_header_style(cell)

    for row_idx, item in enumerate(recurrence_summary, start=2):
        ws.cell(row=row_idx, column=1, value=_safe_str(item.get("sheet_name")))
        ws.cell(row=row_idx, column=2, value=_safe_str(item.get("pattern_label")))
        ws.cell(row=row_idx, column=3, value=item.get("affected_rows"))
        ws.cell(row=row_idx, column=4, value=item.get("affected_pages"))
        ws.cell(row=row_idx, column=5, value="Yes" if item.get("systemic_issue") else "No")
        ws.cell(row=row_idx, column=6, value=_safe_str(item.get("highest_severity")))
        ws.cell(row=row_idx, column=7, value=_safe_str(item.get("representative_insight")))
        ws.cell(row=row_idx, column=8, value=_safe_str(item.get("recommended_action")))
        ws.cell(row=row_idx, column=9, value="\n".join(item.get("example_urls", [])))

    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def enrich_workbook(workbook_path: str, enriched_json_path: str, output_path: str) -> None:
    enriched_payload = _load_enriched_rows(enriched_json_path)
    rows = enriched_payload.get("rows", [])
    recurrence_summary = enriched_payload.get("recurrence_summary", [])

    wb = load_workbook(workbook_path)

    if "Generated Evidence" not in wb.sheetnames:
        raise ValueError("Workbook does not contain a 'Generated Evidence' sheet.")

    ws = wb["Generated Evidence"]
    if _has_existing_ai_columns(ws):
        raise ValueError(
            "Workbook 'Generated Evidence' already contains AI enrichment columns. "
            "Use a clean workbook exported from the checks JSON."
        )

    sheet_data_rows = max(ws.max_row - 1, 0)
    if sheet_data_rows != len(rows):
        raise ValueError(
            f"Row mismatch: workbook has {sheet_data_rows} data rows in 'Generated Evidence' "
            f"but enriched JSON has {len(rows)} rows. Use the same workbook generated from the same checks file."
        )

    _validate_row_alignment(ws, rows)
    _append_ai_columns(ws, rows)
    _write_summary_sheet(wb, recurrence_summary)

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)

    print(f"[OK] AI-enriched workbook written to: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Append AI review columns to workbook.")
    parser.add_argument("--workbook", required=True, help="Path to current exported workbook")
    parser.add_argument("--enriched-json", required=True, help="Path to sheet_checks_ai_enriched.json")
    parser.add_argument("--output", required=True, help="Path to final workbook output")
    args = parser.parse_args()

    enrich_workbook(
        workbook_path=args.workbook,
        enriched_json_path=args.enriched_json,
        output_path=args.output,
    )


if __name__ == "__main__":
    main()
