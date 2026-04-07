import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


SHEET_EXPORT_MAP = {
    "Content": "Content",
    "Labeling": "Labeling",
    "Presentation": "Presentation",
    "Navigation": "Navigation",
    "Interaction": "Interaction",
    "Feedback": "Feedback",
    "Forms": "Forms",
    "Visual hierarchy": "Visual hirarchy",
}

TRUE_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
TRUE_FONT = Font(color="006100", bold=True)

FALSE_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
FALSE_FONT = Font(color="9C0006", bold=True)

NA_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
NA_FONT = Font(color="404040", bold=True)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

LOW_CONF_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
LOW_CONF_FONT = Font(color="7F6000", bold=True)

LINK_FONT = Font(color="0563C1", underline="single")


def _safe_str(value) -> str:
    if value is None:
        return ""
    return str(value)


def load_checks(checks_path: Path) -> dict:
    with checks_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(cell_value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 60)


def normalize_status(raw_status) -> str:
    if raw_status is None:
        return "N/A"

    if isinstance(raw_status, bool):
        return "TRUE" if raw_status else "FALSE"

    value = str(raw_status).strip().upper()

    true_values = {"TRUE", "T", "YES", "Y", "OK", "VALID", "PASS", "PASSED", "1"}
    false_values = {"FALSE", "F", "NO", "N", "X", "FAIL", "FAILED", "0"}
    na_values = {"", "N/A", "NA", "NOT APPLICABLE", "NONE", "NULL", "UNKNOWN"}

    if value in true_values:
        return "TRUE"
    if value in false_values:
        return "FALSE"
    if value in na_values:
        return "N/A"

    return "N/A"


def confidence_band(confidence) -> str:
    try:
        c = float(confidence)
    except Exception:
        return "Unknown"

    if c >= 0.75:
        return "High"
    if c >= 0.45:
        return "Medium"
    return "Low"


def apply_status_style(cell, normalized_status: str) -> None:
    cell.alignment = Alignment(horizontal="center", vertical="center")

    if normalized_status == "TRUE":
        cell.value = "TRUE"
        cell.fill = TRUE_FILL
        cell.font = TRUE_FONT
    elif normalized_status == "FALSE":
        cell.value = "FALSE"
        cell.fill = FALSE_FILL
        cell.font = FALSE_FONT
    else:
        cell.value = "N/A"
        cell.fill = NA_FILL
        cell.font = NA_FONT


def clear_existing_generated_evidence_sheet(wb) -> None:
    if "Generated Evidence" in wb.sheetnames:
        del wb["Generated Evidence"]


def write_answers_to_sheet(ws, sheet_results) -> None:
    for item in sheet_results:
        row = item.get("row")
        normalized_status = normalize_status(item.get("status"))

        if not isinstance(row, int) or row < 1:
            continue

        cell = ws[f"A{row}"]
        apply_status_style(cell, normalized_status)


def _stringify_evidence_row(item: dict) -> str:
    value = item.get("evidence")
    if isinstance(value, list):
        cleaned = [str(x).strip() for x in value if str(x).strip()]
        if cleaned:
            return " | ".join(cleaned)
    elif value is not None and str(value).strip():
        return str(value)

    rationale = _safe_str(item.get("rationale"))
    if rationale:
        return rationale

    return ""


def _source_pages_text(source_pages) -> str:
    if not isinstance(source_pages, list):
        return ""

    parts = []
    for p in source_pages:
        if not isinstance(p, dict):
            continue

        page_name = p.get("page_name", "") or p.get("page_id", "")
        page_url = p.get("page_url", "")
        screenshot_path = p.get("screenshot_path", "")
        parts.append(f"{page_name} | {page_url} | {screenshot_path}")

    return "\n".join(parts)


def build_evidence_sheet(wb, checks_data: dict) -> None:
    clear_existing_generated_evidence_sheet(wb)
    ws = wb.create_sheet("Generated Evidence")

    headers = [
        "Sheet",
        "Row",
        "Criterion",
        "Status",
        "Confidence",
        "Confidence Band",
        "Needs Review",
        "Applicability",
        "Page Name",
        "Page URL",
        "Final URL",
        "Page ID",
        "Screenshot Path",
        "Decision Basis",
        "Rationale",
        "Evidence",
        "Source Pages",
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheets = checks_data.get("sheets", {})
    for sheet_name, sheet_payload in sheets.items():
        for item in sheet_payload.get("results", []):
            normalized_status = normalize_status(item.get("status"))
            confidence_value = item.get("confidence")
            band = item.get("confidence_band") or confidence_band(confidence_value)

            ws.append([
                item.get("sheet", sheet_name),
                item.get("row"),
                item.get("criterion"),
                normalized_status,
                confidence_value,
                band,
                "YES" if item.get("needs_review") else "",
                item.get("applicability", ""),
                item.get("page_name", ""),
                item.get("page_url", ""),
                item.get("final_url", ""),
                item.get("page_id", ""),
                item.get("screenshot_path", ""),
                item.get("decision_basis", ""),
                item.get("rationale", ""),
                _stringify_evidence_row(item),
                _source_pages_text(item.get("source_pages", [])),
            ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        status_cell = row[3]        # D
        conf_band_cell = row[5]     # F
        needs_review_cell = row[6]  # G
        page_url_cell = row[9]      # J
        final_url_cell = row[10]    # K
        screenshot_cell = row[12]   # M

        normalized_status = str(status_cell.value or "").strip().upper()
        if normalized_status == "TRUE":
            status_cell.fill = TRUE_FILL
            status_cell.font = TRUE_FONT
        elif normalized_status == "FALSE":
            status_cell.fill = FALSE_FILL
            status_cell.font = FALSE_FONT
        else:
            status_cell.fill = NA_FILL
            status_cell.font = NA_FONT

        if str(conf_band_cell.value or "").strip() == "Low":
            conf_band_cell.fill = LOW_CONF_FILL
            conf_band_cell.font = LOW_CONF_FONT

        if str(needs_review_cell.value or "").strip().upper() == "YES":
            needs_review_cell.fill = LOW_CONF_FILL
            needs_review_cell.font = LOW_CONF_FONT

        for link_cell in (page_url_cell, final_url_cell):
            if link_cell.value and str(link_cell.value).startswith(("http://", "https://")):
                link_cell.hyperlink = str(link_cell.value)
                link_cell.font = LINK_FONT

        if screenshot_cell.value:
            screenshot_path = str(screenshot_cell.value)
            screenshot_cell.hyperlink = screenshot_path
            screenshot_cell.font = LINK_FONT

    autosize_columns(ws)


def ensure_output_directory(output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)


def save_workbook_safe(wb, output_path: Path) -> None:
    try:
        wb.save(output_path)
    except PermissionError as e:
        raise PermissionError(
            f"Permission denied while saving '{output_path}'. "
            f"Close the Excel file if it is open, or choose a different output filename."
        ) from e


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True, help="Path to the template workbook")
    parser.add_argument("--checks", required=True, help="Path to enriched sheet_checks.json")
    parser.add_argument("--output", required=True, help="Path to output workbook")
    args = parser.parse_args()

    template_path = Path(args.template)
    checks_path = Path(args.checks)
    output_path = Path(args.output)

    if not template_path.exists():
        raise FileNotFoundError(f"Template workbook not found: {template_path}")
    if not checks_path.exists():
        raise FileNotFoundError(f"Checks JSON not found: {checks_path}")

    checks_data = load_checks(checks_path)

    # Helpful debug line: proves the exporter is reading the enriched schema
    first_result = None
    for sheet_payload in checks_data.get("sheets", {}).values():
        results = sheet_payload.get("results", [])
        if results:
            first_result = results[0]
            break
    if first_result:
        print("Loaded result fields:", sorted(first_result.keys()))

    wb = load_workbook(template_path)

    sheets_data = checks_data.get("sheets", {})

    for checks_sheet_name, workbook_sheet_name in SHEET_EXPORT_MAP.items():
        if workbook_sheet_name not in wb.sheetnames:
            print(f"Warning: sheet '{workbook_sheet_name}' not found in workbook, skipped.")
            continue

        ws = wb[workbook_sheet_name]
        sheet_payload = sheets_data.get(checks_sheet_name, {})
        results = sheet_payload.get("results", [])
        write_answers_to_sheet(ws, results)

    build_evidence_sheet(wb, checks_data)

    ensure_output_directory(output_path)
    save_workbook_safe(wb, output_path)

    print(f"Workbook created successfully: {output_path}")


if __name__ == "__main__":
    main()
